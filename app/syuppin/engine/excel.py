import os

from django.db.models.manager import BaseManager
import openpyxl
import shutil

from app.settings import BASE_DIR
from config.const import * 
from setting.models.common_setting import *
from setting.models.syuppin_column_config import *
from syuppin.models.syuppin import *
import ulid


TEMPLATE_SHEET_NAME = "テンプレート"
TARGET_ROW = 3

def create_excel(account_id:str, item_objects:BaseManager[SyuppinItemModel]):
    # テンプレートExcelを作業用にコピー
    work_filename = PATH.EXCEL_WORK_FILE.format(work_file=ulid.new().str) # 重複しないランダムなファイル名
    shutil.copyfile(os.path.join(BASE_DIR, PATH.EXCEL_TEMPLATE), 
                    work_filename)
    
    # 読み込み(マクロファイルのためkeep_vba=True)
    wb = openpyxl.load_workbook(os.path.join(BASE_DIR, work_filename), keep_vba=True)
    #wb = openpyxl.load_workbook(os.path.join(BASE_DIR, PATH.EXCEL_TEMPLATE))
    ws = wb[TEMPLATE_SHEET_NAME]
    
    # 出品データを書き込み
    write_syuppin_data(ws,account_id ,item_objects)
    
    #wb.save(os.path.join(BASE_DIR, "test_xlm.xlsm"))
    
    return wb, work_filename
    
    


def write_syuppin_data(ws,account_id:str, item_objects:BaseManager[SyuppinItemModel]):
    # column_idの列番号を取得
    setting_objects = SyuppinCommonSettingModel.objects.filter(account_id=account_id).order_by('sec_no').all()
    headers = fetch_headers(ws, TARGET_ROW)

    # column_idに対応するitemデータを記述
    #item_objects = SyuppinItemModel.objects.filter(account_id=account_id).all()
    syuppin_column_objects = SyuppinColumnConfigModel.objects.all()
    for row,item_obj in enumerate(item_objects):
        selected_image_count = 0
        # SKUは小文字でないとエラーになるため小文字とする(厳密にユニークな必要はないので末尾１０桁程度とする)
        item_obj.__dict__["item_sku"] = item_obj.__dict__["item_sku"].lower()[-10:]
        for obj in syuppin_column_objects:
            # ExcelのヘッダとDBから取得したヘッダが一致した場合に当該列にデータを出力する
            if headers.get(obj.excel_column_id):
                # 選択した画像のみExcelに出力する
                if obj.syuppin_column_id.find("image") >= 0:
                    image_index = obj.syuppin_column_id[-1]
                    if item_obj.is_selected_image(image_index):
                        ws.cell(row=row+1+TARGET_ROW, column=headers["main_image_url"] + selected_image_count).value = item_obj.__dict__[obj.syuppin_column_id]
                        selected_image_count += 1
                # 画像以外のデータは通常通り出力
                else:
                    ws.cell(row=row+1+TARGET_ROW, column=headers[obj.excel_column_id]).value = item_obj.__dict__[obj.syuppin_column_id]
    # column_idに対応するcommonデータを記述
    for obj in setting_objects:
        if headers.get(obj.column_id):
            for row in range(len(item_objects)):
                ws.cell(row=row+1+TARGET_ROW, column=headers[obj.column_id]).value = obj.column_value


def fetch_headers(ws,target_row:int=TARGET_ROW):
    '''
    Excelのヘッダからカラムを特定するキー情報を取得
    '''
    headers = {}
    for col in range(ws.max_column):
        headers[ws.cell(row=target_row, column=col+1).value] = col + 1
    
    return headers

